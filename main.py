import os
import shutil
import openpyxl


def rename_image_files(root_dir, final_dir):
    """
    Renames the image files based on their directory structure and actual filename.
    Also, collects counts of images in each positive/negative directory.

    Parameters:
    :param root_dir: Path of the root-directory
    :param final_dir: final directory name in which the images are present

    Returns:
    :return: image_count_list containing count of images
    """
    image_count_list = []
    for dir_path, _, filenames in os.walk(root_dir):
        if os.path.basename(dir_path) == final_dir:
            # Counting the images
            image_count = sum(1 for f in filenames if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')))
            image_count_list.append((dir_path, image_count))

            # Breaking down the directory path into subdirectories based on OS
            if os.name == 'nt':
                # Windows
                all_sub_dirs = os.path.normpath(dir_path).split(os.path.sep)
            else:
                # Mac/ Linux
                all_sub_dirs = dir_path.split(os.sep)

            # Checking if we have at least 7 subdirectories to form the new name
            if len(all_sub_dirs) == 7:
                ignored_first = all_sub_dirs[2]  # 575
                site_id = all_sub_dirs[3]  # Site-ID => 575-70-01-0019-01
                camera_id = all_sub_dirs[4]  # Camera-ID => 1
                ignore_date = all_sub_dirs[5]  # Date => 2023-05-29
                ignored_fifth = all_sub_dirs[6]  # positive/ negative
                print(f"Ignoring: {ignored_first}, Saving Site-Id: {site_id}, "
                      f"Appending Camera-ID: {camera_id}, Ignoring Date: {ignore_date}, Ignoring: {ignored_fifth}")

                for file in filenames:
                    if file.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff')):
                        # Getting the filename without extension
                        filename_without_extension = os.path.splitext(file)[0]

                        # Trimming all whitespaces, removing dashes, dots, and the first 2 characters
                        modified_name = filename_without_extension.replace(" ", "").replace("-", "").replace(".", "")
                        modified_name = modified_name[2:]
                        print(f"modified_name = {modified_name}")

                        # Generate new filename using site_id + camera_id + modified_name
                        new_name = f"{site_id}{camera_id}{modified_name}.jpg"
                        new_name = new_name.replace('-', '')
                        print(f"new_name = {new_name}")

                        # Renaming the image file
                        os.rename(os.path.join(dir_path, file), os.path.join(dir_path, new_name))

    return image_count_list


def copy_files_from_directory(root_dir, target_folder, final_dir):
    """
    Copy files from a specific subdirectory name in all directories under root_dir to a target folder.

    Parameters:
    :param root_dir: Path of the root-directory
    :param target_folder: name of the desired folder the images need to be moved into
    :param final_dir: final directory name in which the images are present
    """
    for dir_path, _, filenames in os.walk(root_dir):
        if os.path.basename(dir_path) == final_dir:
            for file in filenames:
                if file.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff')):
                    # Creating target directory if it doesn't exist
                    if not os.path.exists(target_folder):
                        os.makedirs(target_folder)
                    # Copying the image to the target directory
                    shutil.copy2(os.path.join(dir_path, file), target_folder)


def write_to_excel(data, sheet_name, excel_path):
    """
    Writes the number of images in each positive and negative directory per site into Excel

    Parameters:
    :param data:
    :sheet_name: The name of the sheet ("positive" or "negative").
    :param excel_path: path to be saved into
    """
    # Check if the Excel file exists
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
    else:
        wb = openpyxl.Workbook()

    # Check if the sheet name exists
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in data:
            ws.append(row)
    else:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["Directory Path", "Image Count"])
        for row in data:
            ws.append(row)

    # Removing the default sheet if it exists
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    wb.save(excel_path)


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    root_directory = './test'  # Replace with own

    # Directory structure of targets
    target_excel_path = '/Users/ranitrajganguly/Desktop/image_counts.xlsx'  # Replace with own
    target_positive_directory = '/Users/ranitrajganguly/Desktop/thermal_dataset/positive'  # Replace with own
    target_negative_directory = '/Users/ranitrajganguly/Desktop/thermal_dataset/negative'  # Replace with own

    # Renaming positive and negative directories
    positive_renamed_data = rename_image_files(root_directory, 'positive')
    negative_renamed_data = rename_image_files(root_directory, 'negative')

    # Write the counts to Excel
    write_to_excel(positive_renamed_data, 'positive', target_excel_path)
    write_to_excel(negative_renamed_data, 'negative', target_excel_path)

    # Moving renamed images into single file
    copy_files_from_directory(root_directory, target_positive_directory, 'positive')
    copy_files_from_directory(root_directory, target_negative_directory, 'negative')
